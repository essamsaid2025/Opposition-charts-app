"""Automatic file-format detection: encoding, delimiter, sheet names and the
header row - Step 3 of the import wizard."""
from __future__ import annotations

import csv
from dataclasses import dataclass, field
from io import BytesIO

_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
_DELIMITERS = ",;\t|"


@dataclass(frozen=True, slots=True)
class FormatInfo:
    kind: str                          # "csv" | "excel" | "json" | "xml" | "unknown"
    encoding: str = "utf-8"
    delimiter: str = ","
    sheet_names: tuple[str, ...] = ()
    header_row: int = 0
    notes: tuple[str, ...] = field(default_factory=tuple)


def detect_format(data: bytes, filename: str) -> FormatInfo:
    name = filename.lower()
    if name.endswith((".xlsx", ".xls")):
        return _detect_excel(data)
    if name.endswith(".json") or data[:1] in (b"[", b"{"):
        return FormatInfo(kind="json", encoding=_detect_encoding(data))
    if name.endswith(".xml") or data.lstrip()[:1] == b"<":
        return FormatInfo(kind="xml", encoding=_detect_encoding(data))
    return _detect_csv(data)


def _detect_encoding(data: bytes) -> str:
    for enc in _ENCODINGS:
        try:
            data.decode(enc)
            return enc
        except UnicodeDecodeError:
            continue
    return "latin-1"


def _detect_csv(data: bytes) -> FormatInfo:
    encoding = _detect_encoding(data)
    sample = data[:64_000].decode(encoding, errors="replace")
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=_DELIMITERS).delimiter
    except csv.Error:
        counts = {d: sample.count(d) for d in _DELIMITERS}
        delimiter = max(counts, key=counts.get) if any(counts.values()) else ","
    header_row = _guess_header_row(sample.splitlines(), delimiter)
    return FormatInfo(kind="csv", encoding=encoding, delimiter=delimiter, header_row=header_row)


def _guess_header_row(lines: list[str], delimiter: str) -> int:
    """First row where most cells are non-numeric text = header."""
    for i, line in enumerate(lines[:10]):
        cells = [c.strip() for c in line.split(delimiter) if c.strip()]
        if len(cells) >= 2:
            texty = sum(1 for c in cells if not c.replace(".", "", 1).lstrip("-").isdigit())
            if texty / len(cells) > 0.5:
                return i
    return 0


def _detect_excel(data: bytes) -> FormatInfo:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(data), read_only=True)
        sheets = tuple(wb.sheetnames)
        wb.close()
        return FormatInfo(kind="excel", sheet_names=sheets)
    except Exception:
        return FormatInfo(kind="excel", notes=("Could not enumerate sheets",))
