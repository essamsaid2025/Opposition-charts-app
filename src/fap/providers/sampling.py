"""Bounded evidence extraction: read the front of a file, never the whole thing.

Detection must cost the same for a 2 KB export and a million-row one, so every
reader here works on a prefix and stops. The sample is taken once per import
and every provider is scored against that one sample - no provider ever touches
the raw bytes to decide whether it recognizes them.

Format/encoding/delimiter/sheet detection is not re-implemented: it comes from
fap.providers.detection.detect_format, as it always has.
"""
from __future__ import annotations

import csv
import io
import json
import re
from dataclasses import dataclass, field
from typing import Any

from fap.providers.detection import FormatInfo, detect_format

HEAD_BYTES = 64_000          # prefix decoded for text formats
JSON_PARSE_LIMIT = 4_000_000  # above this a JSON file is sampled by regex, not parsed
MAX_SAMPLE_ROWS = 20         # rows read for column/value evidence
MAX_VALUES = 200             # sampled cell values
MAX_NESTED_DEPTH = 3

_KEY_RE = re.compile(r'"([A-Za-z_][A-Za-z0-9_ .-]{0,60})"\s*:')
_NESTED_RE_CACHE: dict[str, re.Pattern[str]] = {}


@dataclass(frozen=True, slots=True)
class FileSample:
    """Everything the intelligence engine is allowed to know about a file."""
    filename: str
    extension: str
    kind: str                                   # csv | excel | json | xml | unknown
    format: FormatInfo
    size_bytes: int
    columns: tuple[str, ...] = ()
    sheet_names: tuple[str, ...] = ()
    metadata: str = ""                          # workbook properties, xml root tag
    json_keys: tuple[str, ...] = ()             # top-level keys
    nested_paths: tuple[str, ...] = ()          # "type.name", "pass.end_location"
    values: tuple[str, ...] = ()
    text_head: str = ""                         # decoded prefix, for fingerprints
    notes: tuple[str, ...] = field(default_factory=tuple)

    def has_nested_path(self, dotted: str) -> bool:
        if dotted in self.nested_paths:
            return True
        # huge files are not parsed; fall back to the decoded prefix
        if not self.text_head:
            return False
        pattern = _NESTED_RE_CACHE.get(dotted)
        if pattern is None:
            parts = [re.escape(p) for p in dotted.split(".")]
            expr = r'"' + parts[0] + r'"\s*:\s*[\[{](?:[^{}\[\]]|\{[^{}]*\})*?"' + parts[-1] + r'"'
            pattern = re.compile(expr, re.DOTALL)
            _NESTED_RE_CACHE[dotted] = pattern
        return bool(pattern.search(self.text_head))


def _decode_head(data: bytes, encoding: str) -> str:
    return data[:HEAD_BYTES].decode(encoding, errors="replace")


def _sample_csv(data: bytes, fmt: FormatInfo) -> dict[str, Any]:
    head = _decode_head(data, fmt.encoding)
    rows = list(csv.reader(io.StringIO(head), delimiter=fmt.delimiter))
    if not rows:
        return {"text_head": head}
    header_row = min(fmt.header_row, len(rows) - 1)
    columns = tuple(c.strip() for c in rows[header_row] if c.strip())
    values: list[str] = []
    for row in rows[header_row + 1: header_row + 1 + MAX_SAMPLE_ROWS]:
        for cell in row:
            cell = cell.strip()
            if cell:
                values.append(cell)
            if len(values) >= MAX_VALUES:
                break
        if len(values) >= MAX_VALUES:
            break
    return {"columns": columns, "values": tuple(values), "text_head": head}


def _sample_excel(data: bytes) -> dict[str, Any]:
    """Header row + sheet names + workbook properties, read-only, no full load."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True)
    except Exception:
        return {"notes": ("workbook could not be opened for sampling",)}
    try:
        sheets = tuple(wb.sheetnames)
        props = wb.properties
        metadata = " ".join(str(x) for x in (
            getattr(props, "creator", "") or "", getattr(props, "title", "") or "",
            getattr(props, "company", "") or "", getattr(props, "description", "") or "",
            getattr(props, "category", "") or "", getattr(props, "keywords", "") or "",
        ) if x).strip()
        sheet = wb[wb.sheetnames[0]]
        columns: tuple[str, ...] = ()
        values: list[str] = []
        for i, row in enumerate(sheet.iter_rows(max_row=MAX_SAMPLE_ROWS + 1, values_only=True)):
            if i == 0:
                columns = tuple(str(c).strip() for c in row if c is not None and str(c).strip())
                continue
            for cell in row:
                if cell is not None and str(cell).strip():
                    values.append(str(cell).strip())
            if len(values) >= MAX_VALUES:
                break
        return {"columns": columns, "sheet_names": sheets, "metadata": metadata,
                "values": tuple(values[:MAX_VALUES])}
    finally:
        wb.close()


def _flatten_paths(obj: Any, prefix: str = "", depth: int = 0) -> list[str]:
    if depth > MAX_NESTED_DEPTH or not isinstance(obj, dict):
        return []
    paths: list[str] = []
    for key, value in obj.items():
        path = f"{prefix}.{key}" if prefix else str(key)
        paths.append(path)
        if isinstance(value, dict):
            paths.extend(_flatten_paths(value, path, depth + 1))
        elif isinstance(value, list) and value and isinstance(value[0], dict):
            paths.extend(_flatten_paths(value[0], path, depth + 1))
    return paths


def _records_of(payload: Any) -> list[Any]:
    if isinstance(payload, list):
        return payload[:MAX_SAMPLE_ROWS]
    if isinstance(payload, dict):
        for value in payload.values():
            if isinstance(value, list) and value and isinstance(value[0], dict):
                return value[:MAX_SAMPLE_ROWS]
        return [payload]
    return []


def _sample_json(data: bytes, fmt: FormatInfo) -> dict[str, Any]:
    head = _decode_head(data, fmt.encoding)
    if len(data) > JSON_PARSE_LIMIT:
        # too big to parse for a guess: the prefix carries enough evidence
        keys = tuple(dict.fromkeys(_KEY_RE.findall(head)))
        return {"json_keys": keys, "columns": keys, "text_head": head,
                "notes": (f"sampled first {HEAD_BYTES // 1000} KB (file is "
                          f"{len(data) // 1_000_000} MB)",)}
    try:
        payload = json.loads(head if head.strip().endswith(("]", "}")) else data.decode(
            fmt.encoding, errors="replace"))
    except Exception:
        try:                                    # JSON Lines: one object per line
            payload = [json.loads(line) for line in head.splitlines()[:MAX_SAMPLE_ROWS]
                       if line.strip()]
        except Exception:
            keys = tuple(dict.fromkeys(_KEY_RE.findall(head)))
            return {"json_keys": keys, "columns": keys, "text_head": head}

    top_keys = tuple(payload.keys()) if isinstance(payload, dict) else ()
    records = _records_of(payload)
    paths: list[str] = []
    columns: list[str] = []
    values: list[str] = []
    for record in records[:MAX_SAMPLE_ROWS]:
        if not isinstance(record, dict):
            continue
        paths.extend(_flatten_paths(record))
        columns.extend(record.keys())
        for value in record.values():
            if isinstance(value, (str, int, float)) and len(values) < MAX_VALUES:
                values.append(str(value))
    return {"json_keys": top_keys or tuple(dict.fromkeys(columns)),
            "columns": tuple(dict.fromkeys(columns)),
            "nested_paths": tuple(dict.fromkeys(paths)),
            "values": tuple(values), "text_head": head}


def _sample_xml(data: bytes, fmt: FormatInfo) -> dict[str, Any]:
    head = _decode_head(data, fmt.encoding)
    tags = tuple(dict.fromkeys(re.findall(r"<([A-Za-z_][\w.-]{0,40})", head)))
    attrs = tuple(dict.fromkeys(re.findall(r'([A-Za-z_][\w.-]{0,40})\s*=\s*"', head)))
    return {"columns": attrs, "json_keys": tags, "metadata": " ".join(tags[:10]),
            "text_head": head}


def sample_file(data: bytes, filename: str) -> FileSample:
    """One bounded pass over the front of a file, shared by every provider."""
    fmt = detect_format(data, filename)
    extension = ("." + filename.rsplit(".", 1)[-1].lower()) if "." in filename else ""
    parts: dict[str, Any] = {}
    if fmt.kind == "excel":
        parts = _sample_excel(data)
    elif fmt.kind == "json":
        parts = _sample_json(data, fmt)
    elif fmt.kind == "xml":
        parts = _sample_xml(data, fmt)
    else:
        parts = _sample_csv(data, fmt)
    return FileSample(filename=filename, extension=extension, kind=fmt.kind, format=fmt,
                      size_bytes=len(data), **parts)
